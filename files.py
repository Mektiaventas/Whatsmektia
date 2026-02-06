import os
import time
import traceback
import zipfile
import shutil
import logging

import pandas as pd
import openpyxl
import PyPDF2
import fitz
from docx import Document
from werkzeug.utils import secure_filename

# Intentar reusar `app` y constantes si el módulo principal ya existe.
try:
    from app import app, UPLOAD_FOLDER, obtener_configuracion_por_host, get_productos_dir_for_config
except Exception:
    # Fallbacks mínimos para desarrollo/local
    app = None
    UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER') or os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'uploads')
    def obtener_configuracion_por_host():
        return {}
    def get_productos_dir_for_config(config=None):
        productos_dir = os.path.join(UPLOAD_FOLDER, 'productos')
        os.makedirs(productos_dir, exist_ok=True)
        return productos_dir, 'default'

# Logger: prefer app.logger cuando esté disponible
logger = app.logger if app is not None else logging.getLogger(__name__)

def determinar_extension(mime_type, filename):
    """Determina la extensión del archivo basado en MIME type y nombre"""
    mime_to_extension = {
        'application/pdf': 'pdf',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'xlsx',
        'application/vnd.ms-excel': 'xls',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'docx',
        'text/csv': 'csv',
        'text/plain': 'txt'
    }
    
    # Primero intentar por MIME type
    extension = mime_to_extension.get(mime_type)
    
    # Si no se encuentra, intentar por extensión del nombre de archivo
    if not extension and '.' in filename:
        extension = filename.split('.')[-1].lower()
    
    return extension or 'bin'

def extraer_imagenes_embedded_excel(filepath, output_dir=None, config=None):
    """
    Versión CORREGIDA: Extrae imágenes vinculándolas estrictamente a su fila.
    """
    try:
        if output_dir is None:
            try:
                productos_dir, tenant_slug = get_productos_dir_for_config(config)
                output_dir = productos_dir
            except Exception as e:
                logger.warning(f"⚠️ get_productos_dir_for_config falló, usando legacy. Error: {e}")
                output_dir = os.path.join(UPLOAD_FOLDER, 'productos')

        os.makedirs(output_dir, exist_ok=True)

        wb = openpyxl.load_workbook(filepath)
        imagenes_extraidas = []

        for sheet in wb.worksheets:
            # Importante: openpyxl guarda las imágenes en _images
            for idx, img in enumerate(getattr(sheet, '_images', [])):
                try:
                    # 1. Extraer la imagen real
                    img_obj = img.image
                    img_format = (img_obj.format or 'PNG').lower()
                    
                    # 2. Determinar posición (Ancla)
                    row = None
                    col = None
                    
                    # Intentar obtener la fila de la forma más directa posible
                    try:
                        if hasattr(img.anchor, '_from'):
                            row = img.anchor._from.row + 1
                            col = img.anchor._from.col + 1
                        elif hasattr(img.anchor, 'row'): # Para otros tipos de ancla
                            row = img.anchor.row
                            col = img.anchor.col
                    except:
                        pass

                    # Si no pudimos obtener la fila, no nos sirve para el catálogo
                    if row is None:
                        logger.warning(f"⏩ Imagen {idx} en {sheet.title} saltada: no se detectó fila (flotante).")
                        continue

                    # 3. Guardar con nombre único incluyendo la fila para evitar confusiones
                    timestamp = int(time.time())
                    img_filename = f"excel_fila_{row}_{idx+1}_{timestamp}.{img_format}"
                    img_path = os.path.join(output_dir, img_filename)

                    img_obj.save(img_path)

                    imagenes_extraidas.append({
                        'filename': img_filename,
                        'path': img_path,
                        'sheet': sheet.title,
                        'anchor': str(img.anchor),
                        'row': row,
                        'col': col
                    })
                    logger.info(f"✅ Imagen vinculada: {img_filename} -> FILA {row}")

                except Exception as e:
                    logger.warning(f"⚠️ Error en imagen {idx}: {e}")
                    continue

        return imagenes_extraidas

    except Exception as e:
        logger.error(f"🔴 Error crítico en extraer_imagenes_embedded_excel: {e}")
        return []

def _extraer_imagenes_desde_zip_xlsx(filepath, output_dir):
    """
    Fallback CORREGIDO: Extrae imágenes ordenándolas numéricamente para evitar desfaces.
    """
    import re # Necesario para el ordenamiento natural
    os.makedirs(output_dir, exist_ok=True)
    imagenes = []
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            # 1. Filtramos los archivos de media
            media_files = [f for f in z.namelist() if f.startswith('xl/media/')]
            
            # 2. ORDENAMIENTO NATURAL: Esto asegura que image2 esté antes que image10
            # y que image289 esté antes que image290
            media_files.sort(key=lambda x: [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', x)])
            
            for idx, media_path in enumerate(media_files):
                try:
                    ext = os.path.splitext(media_path)[1].lstrip('.').lower() or 'bin'
                    timestamp = int(time.time())
                    
                    # Usamos el nombre original del zip para debuguear si es necesario
                    # pero mantenemos tu estructura de nombre excel_unzip_img_...
                    filename = f"excel_unzip_img_{idx+1}_{timestamp}.{ext}"
                    dest_path = os.path.join(output_dir, filename)
                    
                    with z.open(media_path) as src, open(dest_path, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
                    
                    imagenes.append({
                        'filename': filename,
                        'path': dest_path,
                        'sheet': None,
                        'anchor': None,
                        'row': None,
                        'col': None,
                        'original_zip_name': media_path # Agregamos esto para saber cuál era en el ZIP
                    })
                    logger.info(f"✅ Imagen (zip) extraída: {filename} desde {media_path}")
                except Exception as e:
                    logger.warning(f"⚠️ No se pudo extraer {media_path} desde zip: {e}")
    except zipfile.BadZipFile:
        logger.warning("⚠️ Archivo no es un .xlsx válido o está corrupto; zip fallback falló")
    except Exception as e:
        logger.warning(f"⚠️ Error extrayendo imágenes desde zip: {e}")
    return imagenes

def get_docs_dir_for_config(config=None):
    """Return (docs_dir, tenant_slug). Ensures uploads/docs/<tenant_slug> exists.

    Prefer the runtime Flask UPLOAD_FOLDER (current_app or imported app.config) to avoid
    import-time / env var races that cause files to be written to a different base path.
    """
    # Prefer runtime app.config['UPLOAD_FOLDER'] when possible
    base_upload = None
    try:
        from flask import has_request_context, current_app
        if has_request_context():
            base_upload = current_app.config.get('UPLOAD_FOLDER')
    except Exception:
        # no request context or flask not available here
        pass

    # If not available via current_app, try module-level app (import-time)
    if not base_upload:
        try:
            if 'app' in globals() and getattr(app, 'config', None):
                base_upload = app.config.get('UPLOAD_FOLDER')
        except Exception:
            base_upload = None

    # Fallback to environment or module-level constant
    if not base_upload:
        base_upload = os.environ.get('UPLOAD_FOLDER') or UPLOAD_FOLDER

    if config is None:
        try:
            config = obtener_configuracion_por_host()
        except Exception:
            config = {}

    dominio = (config.get('dominio') or '').strip().lower()
    tenant_slug = dominio.split('.')[0] if dominio else 'default'

    docs_dir = os.path.join(os.path.abspath(base_upload), 'docs', tenant_slug)
    try:
        os.makedirs(docs_dir, exist_ok=True)
    except Exception as e:
        logger.warning(f"⚠️ No se pudo crear docs_dir {docs_dir}: {e}")
        # fallback to shared docs dir under same base
        docs_dir = os.path.join(os.path.abspath(base_upload), 'docs')
        os.makedirs(docs_dir, exist_ok=True)

    return docs_dir, tenant_slug

def get_productos_dir_for_config(config=None):
    """Return (productos_dir, tenant_slug). Ensures uploads/productos/<tenant_slug> exists."""
    if config is None:
        try:
            config = obtener_configuracion_por_host()
        except Exception:
            config = {}
    dominio = (config.get('dominio') or '').strip().lower()
    tenant_slug = dominio.split('.')[0] if dominio else 'default'
    productos_dir = os.path.join(os.path.abspath(UPLOAD_FOLDER), 'productos', tenant_slug)
    os.makedirs(productos_dir, exist_ok=True)
    return productos_dir, tenant_slug

def extraer_texto_e_imagenes_pdf(file_path):
    """Extrae texto e imágenes de un archivo PDF"""
    try:
        texto = ""
        imagenes = []
        
        # Abrir el PDF con PyMuPDF
        doc = fitz.open(file_path)
        
        # Crear directorio para imágenes si no existe (tenant-aware fallback)
        try:
            productos_dir, tenant_slug = get_productos_dir_for_config()
            img_dir = productos_dir
        except Exception as e:
            logger.warning(f"⚠️ get_productos_dir_for_config falló, usando legacy uploads/productos. Error: {e}")
            img_dir = os.path.join(UPLOAD_FOLDER, 'productos')
        os.makedirs(img_dir, exist_ok=True)
        
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            
            # Extraer texto
            texto += page.get_text()
            
            # Extraer imágenes
            image_list = page.get_images(full=True)
            for img_idx, img_info in enumerate(image_list):
                try:
                    xref = img_info[0]
                    
                    # Verificar si la imagen es válida antes de procesarla
                    try:
                        base_img = doc.extract_image(xref)
                        
                        # Obtener la imagen en bytes
                        imagen_bytes = base_img["image"]
                        
                        # Determinar formato de imagen
                        extension = base_img["ext"]
                        
                        # Crear nombre único para la imagen
                        img_filename = f"producto_{page_num+1}_{img_idx+1}_{int(time.time())}.{extension}"
                        img_path = os.path.join(img_dir, img_filename)
                        
                        # Guardar la imagen
                        with open(img_path, "wb") as img_file:
                            img_file.write(imagen_bytes)
                        
                        # Intentar obtener el rectángulo de la imagen de manera segura
                        try:
                            rect = page.get_image_bbox(xref)
                        except ValueError:
                            # Si falla, usar un rectángulo vacío
                            rect = fitz.Rect(0, 0, 0, 0)
                        
                        # Agregar a la lista de imágenes con metadatos útiles
                        imagenes.append({
                            'filename': img_filename,
                            'path': img_path,
                            'page': page_num,
                            'size': len(imagen_bytes),
                            'position': img_info[1:],  # Info de posición para asociar con texto
                            'xref': xref,
                            'rect': rect
                        })
                        
                        logger.info(f"✅ Imagen extraída: {img_filename} (tenant_dir={img_dir})")
                        
                    except Exception as e:
                        logger.warning(f"⚠️ Error extrayendo imagen específica {xref}: {e}")
                        continue
                        
                except Exception as e:
                    logger.warning(f"⚠️ Error procesando imagen {img_idx} en página {page_num+1}: {e}")
                    continue
        
        doc.close()
        
        logger.info(f"✅ Texto extraído: {len(texto)} caracteres")
        logger.info(f"🖼️ Imágenes extraídas: {len(imagenes)}")
        
        return texto.strip(), imagenes
        
    except Exception as e:
        logger.error(f"🔴 Error extrayendo contenido PDF: {e}")
        logger.error(traceback.format_exc())
        
        # Intenta al menos extraer el texto usando el método anterior
        try:
            texto = extraer_texto_pdf(file_path)
            logger.info(f"✅ Se pudo extraer texto con método alternativo: {len(texto)} caracteres")
            return texto, []  # Devolver texto pero sin imágenes
        except:
            return None, []

def extraer_texto_pdf(file_path):
    """Extrae texto de un archivo PDF"""
    try:
        texto = ""
        
        # Intentar con PyMuPDF primero (más robusto)
        try:
            doc = fitz.open(file_path)
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                texto += page.get_text()
            doc.close()
            logger.info(f"✅ Texto extraído con PyMuPDF: {len(texto)} caracteres")
            return texto.strip()
        except Exception as e:
            logger.warning(f"⚠️ PyMuPDF falló, intentando con PyPDF2: {e}")
        
        # Fallback a PyPDF2
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                texto += page.extract_text()
        
        logger.info(f"✅ Texto extraído con PyPDF2: {len(texto)} caracteres")
        return texto.strip()
        
    except Exception as e:
        logger.error(f"🔴 Error extrayendo texto PDF: {e}")
        return None

def extraer_texto_excel(filepath):
    """Extrae texto de archivos Excel"""
    try:
        texto = ""
        
        # Leer todas las hojas
        if filepath.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filepath)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                texto += f"\n--- Hoja: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    fila_texto = " | ".join(str(cell) for cell in row if cell is not None)
                    if fila_texto.strip():
                        texto += fila_texto + "\n"
        
        # Alternativa con pandas para mejor compatibilidad
        try:
            excel_file = pd.ExcelFile(filepath)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                texto += f"\n--- Hoja: {sheet_name} (Pandas) ---\n"
                texto += df.to_string() + "\n"
        except Exception as e:
            logger.warning(f"⚠️ Pandas falló: {e}")
        
        return texto.strip() if texto.strip() else None
        
    except Exception as e:
        logger.error(f"🔴 Error procesando Excel: {e}")
        return None

def extraer_texto_csv(filepath):
    """Extrae texto de archivos CSV"""
    try:
        df = pd.read_csv(filepath)
        return df.to_string()
    except Exception as e:
        logger.error(f"🔴 Error leyendo CSV: {e}")
        # Intentar lectura simple
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
        except:
            return None

def extraer_texto_docx(filepath):
    """Extrae texto de archivos Word"""
    try:
        doc = Document(filepath)
        texto = ""
        for paragraph in doc.paragraphs:
            texto += paragraph.text + "\n"
        return texto.strip() if texto.strip() else None
    except Exception as e:
        logger.error(f"🔴 Error leyendo DOCX: {e}")
        return None

def extraer_texto_archivo(filepath, extension):
    """Extrae texto de diferentes tipos de archivos"""
    try:
        if extension == 'pdf':
            return extraer_texto_pdf(filepath)
        
        elif extension in ['xlsx', 'xls']:
            return extraer_texto_excel(filepath)
        
        elif extension == 'csv':
            return extraer_texto_csv(filepath)
        
        elif extension == 'docx':
            return extraer_texto_docx(filepath)
        
        elif extension == 'txt':
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
        
        else:
            logger.warning(f"⚠️ Formato no soportado: {extension}")
            return None
            
    except Exception as e:
        logger.error(f"🔴 Error extrayendo texto de {extension}: {e}")
        return None
